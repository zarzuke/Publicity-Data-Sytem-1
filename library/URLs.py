import sqlite3 
from flask import render_template, request,redirect,url_for,flash,session,g,send_file
from library.DB import *
from openpyxl import load_workbook
import io
import os
from datetime import datetime, timedelta
import tempfile
import win32com.client as win32
import pythoncom 

def respaldar_base_de_datos():
    ruta_original = "library/database.db"
    carpeta_destino = "library/database_backrest"

    if not os.path.exists(carpeta_destino):
        os.makedirs(carpeta_destino)
    
    nombre_archivo = os.path.basename(ruta_original)
    ruta_destino = os.path.join(carpeta_destino, nombre_archivo)
    
    with open(ruta_original, 'rb') as archivo_origen:
        contenido = archivo_origen.read()
    
    with open(ruta_destino, 'wb') as archivo_destino:
        archivo_destino.write(contenido)

def try_signup():
    workers=get_users()
    if request.method=="POST":
        bd = sqlite3.connect("library/database.db")
        cursor = bd.cursor()
        username=request.form["username"]
        user_confirmation = user_exists(username)
        if user_confirmation:
            flash("El usuario ya se encuentra registrado en el sistema.")
        else:
            password=request.form["password"]
            role=request.form["role"]
            cursor.execute("INSERT INTO users (userName,userPassword,userRol) VALUES (?, ?, ?)", (username, password, role,))
        bd.commit()
        bd.close()    
    return render_template("settings-user.html",user=g.user,workers=workers)

def try_login():
    if g.user==["vacio","vacio"]:
        render_template("index.html")
    else:
        home=try_home()
        return home
   
    if request.method=="POST":
        bd = sqlite3.connect("library/database.db")
        cursor = bd.cursor()
        username=request.form["username"]
        password=(request.form["password"])
        cursor.execute("SELECT userPassword FROM users WHERE userName = ?" , (username,))
        pw=cursor.fetchone()
        if pw and pw[0]== password:
            cursor.execute("SELECT userRol FROM users WHERE userName = ?" , (username,))
            rol=cursor.fetchone()
            session["username"]= [username,rol[0]]
            return redirect("/home")
        else:
                flash("Credenciales inválidas, intente de nuevo")
    return render_template("index.html")

def try_home():
    match g.user[1]:
        case 'Administrator':
            filas,tipos=get_project()
            names,numbers=get_clients01()
            design,craft,install=get_workers()
            clients=zip(names,numbers)
            combine=zip(filas,tipos)
            return render_template("home.html",user=g.user,filas=combine,clients=clients,design=design,craft=craft,install=install)
        case 'Designer':
            return redirect(url_for("design", designer=g.user[1]))
        case 'Crafter':
            return redirect(url_for("crafting"))
        case 'Installer':
            return redirect(url_for("ending"))
        case "vacio":
            flash("Debe de iniciar sesión primero.")
            return redirect(url_for('login'))

def try_design():
    if g.user==["vacio","vacio"]:
        flash("Debe de iniciar sesión primero.")
        return redirect(url_for('login'))
    else:
        
        if g.user[1]=="Administrator":
            estado = 'Diseño'
            filas,tipos=get_project_phase(1)
            combine=zip(filas,tipos)
            return render_template("design.html",user=g.user,filas=combine,estado=estado)
        else:
            estado = 'Diseño'
            worker=g.user[0]
            filas,tipos=get_project_designer(worker)
            combine=zip(filas,tipos)
            return render_template("design.html",user=g.user,filas=combine,estado=estado)

    
def try_approval():
    if g.user==["vacio","vacio"]:
        flash("Debe de iniciar sesión primero.")
        return redirect(url_for('login'))
    else:
        estado = 'Aprobación'
        filas,tipos=get_project_phase(2)
        combine=zip(filas,tipos)
        return render_template("design.html",user=g.user,filas=combine,estado=estado)

def try_crafting():
    if g.user==["vacio","vacio"]:
        flash("Debe de iniciar sesión primero.")
        return redirect(url_for('login'))
    else:
        if g.user[1]=="Administrator":
            estado = 'Creacion'
            filas,tipos=get_project_phase(3)
            combine=zip(filas,tipos)
            return render_template("design.html",user=g.user,filas=combine,estado=estado)
        else:
            estado = 'Creación'
            worker=g.user[0]
            filas,tipos=get_project_crafter(worker)
            combine=zip(filas,tipos)
            return render_template("design.html",user=g.user,filas=combine,estado=estado)

def try_ending():
    if g.user==["vacio","vacio"]:
        flash("Debe de iniciar sesión primero.")
        return redirect(url_for('login'))
    else:
        if g.user[1]=="Administrator":
            estado = 'Entrega'
            filas,tipos=get_project_phase(4)
            combine=zip(filas,tipos)
            return render_template("design.html",user=g.user,filas=combine,estado=estado)
        else:
            estado = 'Entrega'
            worker=g.user[0]
            filas,tipos=get_project_installer(worker)
            combine=zip(filas,tipos)
            return render_template("design.html",user=g.user,filas=combine,estado=estado)
    
def try_logout():
    session.pop("username", None)
    flash("Sesión cerrada correctamente.")
    return redirect(url_for('login'))

def try_clients():
    if g.user==["vacio","vacio"]:
        flash("Debe de iniciar sesión primero.")
        return redirect(url_for('login'))
    else:
        cliente=get_clients()
        return render_template("clients.html",user=g.user,cliente=cliente,os=os)


def try_form():
    form=project_inc()
    return form

def try_work(id):
    
    if g.user==["vacio","vacio"]:
        flash("Debe de iniciar sesión primero.")
        return redirect(url_for('login'))
    else:
        check=check_id(id)
        if check==1:
            workers=get_workers_id(id)
            if g.user[1]=='Administrator':
                work,types,id=get_details(id)
                return render_template("work.html",user=g.user,details=work,types=types,nor=id,os=os)
            elif g.user[0] in workers[0]:
                work,types,id=get_details(id)
                return render_template("work.html",user=g.user,details=work,types=types,nor=id,os=os)
            else:
                flash("No tienes permiso para acceder a este trabajo.")
                return redirect('/home')
        else:
            flash("El trabajo no existe")
            return redirect('/home')

def try_client(client):
    if g.user==["vacio","vacio"]:
        flash("Debe de iniciar sesión primero.")
        return redirect(url_for('login'))
    
    elif g.user[1]!="Administrator":
        flash("No tienes permisos para acceder a este módulo.")
        return redirect("/home")
    
    else:
        work,types=get_works_client(client)
        combine=zip(work,types)
        return render_template("design.html",user=g.user,filas=combine)

def try_settings():
    if g.user==["vacio","vacio"]:
        flash("Debe de iniciar sesión primero.")
        return redirect(url_for('login'))
    
    elif g.user[1]!="Administrator":
        flash("No tienes permisos para acceder a este módulo.")
        return redirect("/home")
    
    else:
        return render_template("settings.html",user=g.user)

def try_record():
    if g.user==["vacio","vacio"]:
        flash("Debe de iniciar sesión primero.")
        return redirect(url_for('login'))
    
    elif g.user[1]!="Administrator":
        flash("No tienes permisos para acceder a este módulo.")
        return redirect("/home")
    
    else:
        filas = get_record()
        names,numbers=get_clients01()
        design,craft,install=get_workers()
        clients=zip(names,numbers)
        return render_template("settings-record.html",user=g.user,filas=filas,clients=clients)
    
def try_record_fildered(client,date):
    if g.user==["vacio","vacio"]:
        flash("Debe de iniciar sesión primero.")
        return redirect(url_for('login'))
    elif g.user[1]!="Administrator":
        flash("No tienes permisos para acceder a este módulo.")
        return redirect("/home")
    else:
        filas = record(client,date)
        cabecera = ['Nombre', 'Cliente', 'Precio', 'Divisa', 'Fecha', 'Tipo de Trabajo','Trabajadores','Nombre del Archivo']
        filas.insert(0, cabecera)
        return filas

def save_excel_as_pdf(excel_file, pdf_output_file):
    # Inicializar la aplicación de Excel
    pythoncom.CoInitialize()  # Inicializar COM
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False  # No mostrar la aplicación de Excel
    try:
        # Abrir el archivo de Excel
        wb = excel.Workbooks.Open(excel_file)
        # Guardar como PDF
        wb.ExportAsFixedFormat(0, pdf_output_file)  # 0 es para PDF
    finally:
        wb.Close(False)
        excel.Quit()

def try_record_file(client, date, format):
    download_name = 'all_record'
    if date:
        año, semana = date.split('-W')
        año = int(año)
        semana = int(semana)
        primer_dia_semana = datetime.fromisocalendar(año, semana, 1).date()
        ultimo_dia_semana = primer_dia_semana + timedelta(days=6)
        download_name = f"[{primer_dia_semana}]-[{ultimo_dia_semana}]"

    filas = record(client, date)
    try:
        # Cargar la plantilla de Excel
        wb = load_workbook('library/Plantilla.xlsx')
        ws = wb.active

        # Establecer el nombre de descarga si existe la fecha
        if date:
            ws.cell(row=1, column=6, value=download_name)

        # Agregar datos a la hoja de Excel
        start_row = 3
        for i, fila in enumerate(filas, start=start_row):
            for j, value in enumerate(fila, start=1):
                ws.cell(row=i, column=j, value=value)

        # Crear un archivo temporal para Excel
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as excel_temp_file:
            wb.save(excel_temp_file.name)  # Guardar el Excel temporalmente
            excel_temp_path = excel_temp_file.name  # Guardar la ruta para luego convertir a PDF

        # Definir la ruta de salida del PDF
        pdf_output_file = os.path.join(tempfile.gettempdir(), f"{download_name}.pdf")
        
        # Guardar el archivo de Excel como PDF
        save_excel_as_pdf(excel_temp_path, pdf_output_file)

    except Exception as e:
        return f"Error al procesar el archivo: {e}"

    # Enviar el archivo PDF para que el cliente lo descargue
    if format == "pdf":
        return send_file(pdf_output_file, as_attachment=True, download_name=f"{download_name}.pdf", mimetype='application/pdf')

    # Enviar el archivo Excel si se solicita en formato .xlsx
    if format == "xlsx":
        with open(excel_temp_path, 'rb') as excel_file:
            buffer = io.BytesIO(excel_file.read())
            buffer.seek(0)
            return send_file(buffer, as_attachment=True, download_name=f"{download_name}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def try_create_client(client_name):
    client_confirmation = client_exists(client_name)
    if client_confirmation:
        flash("El cliente ya se encuentra registrado en el sistema.")
    else:
        country = request.form.get("country-code", "")
        phone = request.form.get("phone", "")
        if phone:
            number = f"{country} {phone}".strip()
            
        create_client(client_name,number)
        flash("Cliente creado correctamente.")
    return redirect("/settings/client")
    
def try_edit_clients():
    if g.user==["vacio","vacio"]:
        flash("Debe de iniciar sesión primero.")
        return redirect(url_for('login'))
    else:
        cliente=get_clients()
        return render_template("settings-client.html",user=g.user,cliente=cliente,os=os)

def try_delete_client(id):
    name=get_client_name(id)
    text=f"{name} Eliminado de la lista de clientes."
    insertar_notificacion(text,g.user[0])
    delete_client(id)
    flash("Cliente borrado correctamente.")
    return redirect("/settings/client")

def try_delete(id):
    name=get_titulo(id)
    text=f"{name} Eliminado de la lista de trabajos."
    insertar_notificacion(text,g.user[0])
    delete_projects(id)
    flash("Trabajo eliminado correctamente.")
    return redirect("/home")
    
def try_comments(user,worker,comments):
    update(user,worker,comments)
    return redirect(url_for("work",user=user))

def try_open(id, title, client):
    folder_path = f"trabajos\\[{client}]-{title}"
    try:
        os.startfile(folder_path)  # Para Windows
        return f""
    except Exception as e:
        return f"Error en las carpetas"
    
    
def try_next(id,status,cdr):
    phase=check_phase(id)
    if status=="none":
        flash('Datos incorrectos, seleccione estatus.')
        return redirect(url_for('work', user=id))
    
    if status=="canceled":
        if phase[0][0]==1:
            flash('Opción incorrecta.')
            if g.user[1]=="Administrator":
                return redirect(url_for('work', user=id))
            else:
                return redirect("/home")
        else:
            if phase[0][0]==2:
                return_phase(id)
                name=get_titulo(id)
                worker=get_workers_id(id)
                text=f"{name} Diseño negado, Regresando a DISEÑADOR.@{worker[0][0]} revisa las indicaciones de Correccion"
                insertar_notificacion(text,g.user[0])
                flash('Fase Negada')
                return redirect(url_for('work', user=id))
            else:
                return_phase(id)
                flash('Fase Negada')
                return redirect(url_for('work', user=id))
    
    if phase[0][0]==4:
        save_record(id)
        name=get_titulo(id)
        delete_projects(id)
        text=f"{name} - FINALIZADO, ¡buen trabajo equipo!"
        insertar_notificacion(text,g.user[0])
        flash('Trabajo finalizado.')
        return redirect("/home")
    elif phase[0][0]==1:
        if status=="approved":
            change_phase(id)
            add_file(id,cdr)
            name=get_titulo(id)
            text=f"{name} - DISEÑO finalizado, esperando por su APROBACIÓN."
            insertar_notificacion(text,g.user[0])
            flash('Fase APROBADA.')
            if g.user[1]=="Administrator":
                return redirect(url_for('work', user=id))
            else:
                return redirect("/home")
    else:
        if status=="approved":
            change_phase(id)
            if phase[0][0]==2:
                name=get_titulo(id)
                text=f"{name} - APROBACIÓN confirmada, esperando por ELABORACIÓN."
                insertar_notificacion(text,g.user[0])
            if phase[0][0]==3:
                name=get_titulo(id)
                text=f"{name} - ELABORACIÓN completada, esperando por INSTALACIÓN."
                insertar_notificacion(text,g.user[0])
            flash('Fase APROBADA.')
            if g.user[1]=="Administrator":
                return redirect(url_for('work', user=id))
            else:
                return redirect("/home")

def try_update_balance(id,charge,comments):
    if charge=="":
        flash("El valor ingesado es incorrecto, por favor verifique e intente de nuevo.")
        return redirect(url_for('work', user=id))
    else:
        update_currency(id,charge)
        text=f"Cambio de costo Total a {charge} por motivo de: {comments}"
        update(id,"Sistema",text)
    
def try_down_payment(id,down):
    if down=="":
        flash("El valor ingresado es incorrecto, por favor verifique e intente de nuevo.")
        return redirect(url_for('work', user=id))
    else:
        total=get_total(id)
        if float(down)>float(total):
            flash("El abono no puede ser SUPERIOR al COSTO TOTAL del trabajo.")
            return redirect(url_for('work', user=id))
        else:
            down=float(down)
            update_down(id,down)
            text=f"El cliente hizo un abono de: {down}"
            update(id,"Sistema",text)
    
def try_edit_client(id):
    function = edit_client_list(id)
    return function

def try_edit_user():
    function = edit_client()
    return function

def try_delete_user(username):
    name=username
    text=f"{name} Eliminado de la lista de usuarios."
    insertar_notificacion(text,g.user[0])
    function = delete_user(username)
    return function

def try_show_notifications():
    return render_template("notifications.html",user=g.user)